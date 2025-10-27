import requests
from requests.auth import HTTPBasicAuth
import openpyxl
from openpyxl.utils import get_column_letter
import os
import time

# Jira Configuration
jira_url = os.getenv("JIRA_URL", "https://techcarrot-team-aqqopo6gxdmd.atlassian.net")
jira_email = os.getenv("JIRA_EMAIL")
jira_api_token = os.getenv("JIRA_API_TOKEN")
jira_project_key = os.getenv("JIRA_PROJECT_KEY", "PROJ")
jira_label_filter = os.getenv("JIRA_LABEL_FILTER", "")

STATE_MAPPING = {
    "Open": "New",
    "New": "New",
    "To Do": "New",
    "Reopen": "Reopen",
    "In Progress": "New",
    "In Development": "New",
    "Done": "Closed",
    "Closed": "Closed",
    "Resolved": "Resolved"
}

print("🔄 Starting Jira defects extraction...")
start_time = time.time()

# Build JQL
if jira_label_filter:
    if ' ' in jira_project_key:
        jql_query = f'project = "{jira_project_key}" AND type = Bug AND labels = "{jira_label_filter}" ORDER BY created DESC'
    else:
        jql_query = f'project = {jira_project_key} AND type = Bug AND labels = "{jira_label_filter}" ORDER BY created DESC'
    print(f"🏷️ Using label filter: {jira_label_filter}")
else:
    if ' ' in jira_project_key:
        jql_query = f'project = "{jira_project_key}" AND type = Bug ORDER BY created DESC'
    else:
        jql_query = f'project = {jira_project_key} AND type = Bug ORDER BY created DESC'

print(f"📋 Fetching issues from Jira project: {jira_project_key}")
print(f"🔍 JQL Query: {jql_query}")

search_url = f"{jira_url}/rest/api/3/search/jql"
auth = HTTPBasicAuth(jira_email, jira_api_token)
all_issues = []
next_page_token = None
max_results = 100

# Fetch issues
while True:
    params = {"jql": jql_query, "maxResults": max_results, "fields": "*all"}
    if next_page_token:
        params["nextPageToken"] = next_page_token
    try:
        response = requests.get(search_url, params=params, auth=auth, timeout=30)
        if response.status_code != 200:
            print(f"❌ Error fetching issues: {response.status_code} - {response.text}")
            break
        data = response.json()
        issues = data.get("issues", [])
        all_issues.extend(issues)
        total = data.get("total", len(all_issues))
        print(f"   Fetched {len(all_issues)}/{total} issues...")
        next_page_token = data.get("nextPageToken")
        if data.get("isLast", True) or not next_page_token:
            break
    except Exception as e:
        print(f"⚠️ Error: {str(e)}")
        break

if not all_issues:
    print("⚠️ No issues found.")
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Jira Defects"
    sheet.append(["ID", "Work Item Type", "Title", "State", "Original Jira State", "Assigned To", "Tags", "Severity", "Issue Links"])
    current_dir = os.path.dirname(os.path.abspath(__file__))
    data_folder = os.path.join(current_dir, "data")
    os.makedirs(data_folder, exist_ok=True)
    save_path = os.path.join(data_folder, f"Jira {jira_project_key} Defects.xlsx")
    wb.save(save_path)
    print(f"✅ Empty Excel file created at: {save_path}")
    exit()

print(f"✅ Found {len(all_issues)} issues. Processing...")

# 🔍 Step 1: Detect 'Severity' field ID from Jira metadata
severity_field_key = None
fields_url = f"{jira_url}/rest/api/3/field"
print("\n🔍 Checking Jira fields metadata for 'Severity' field...")
try:
    fields_resp = requests.get(fields_url, auth=auth)
    if fields_resp.status_code == 200:
        for f in fields_resp.json():
            if "severity" in f["name"].lower():
                severity_field_key = f["id"]
                print(f"✅ Found Severity field: {f['name']} → {f['id']}")
                break
    else:
        print(f"⚠️ Failed to fetch Jira fields metadata: {fields_resp.status_code}")
except Exception as e:
    print(f"⚠️ Error checking fields metadata: {str(e)}")

if not severity_field_key:
    print("⚠️ Could not detect Severity field in metadata, will check issue fields directly.")

# Create Excel
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Jira Defects"
sheet.append(["ID", "Work Item Type", "Title", "State", "Original Jira State", "Assigned To", "Tags", "Severity", "Issue Links"])

# Process issues
for idx, issue in enumerate(all_issues, start=2):
    fields = issue.get("fields", {})
    issue_key = issue.get("key", "")

    issue_type_obj = fields.get("issuetype")
    issue_type = issue_type_obj.get("name", "Bug") if isinstance(issue_type_obj, dict) else "Bug"

    title = fields.get("summary") or issue_key
    if isinstance(title, dict):
        title = title.get("content", "") or str(title)

    status_obj = fields.get("status")
    jira_status = status_obj.get("name", "Unknown") if isinstance(status_obj, dict) else "Unknown"
    mapped_state = STATE_MAPPING.get(jira_status, jira_status)

    assignee_obj = fields.get("assignee")
    if isinstance(assignee_obj, dict):
        assignee_name = (
            assignee_obj.get("displayName") or
            assignee_obj.get("emailAddress", "").split("@")[0] or
            "Unassigned"
        )
    else:
        assignee_name = "Unassigned"

    # === Severity Extraction ===
    severity_value = None

    # Attempt 1: use detected Severity field ID from metadata
    if severity_field_key:
        severity_obj = fields.get(severity_field_key)
        if isinstance(severity_obj, dict):
            severity_value = severity_obj.get("value") or severity_obj.get("name")
        elif isinstance(severity_obj, str):
            severity_value = severity_obj

    # Attempt 2: common Severity field names
    if not severity_value:
        for field_name in ["Severity", "severity", "SEVERITY"]:
            val = fields.get(field_name)
            if val:
                if isinstance(val, dict):
                    severity_value = val.get("value") or val.get("name")
                elif isinstance(val, str):
                    severity_value = val
                break

    # Attempt 3: Fallback to Priority (only if absolutely needed)
    if not severity_value:
        priority_obj = fields.get("priority")
        if isinstance(priority_obj, dict):
            severity_value = priority_obj.get("name", "Medium")
            if idx <= 3:
                print(f"   ℹ️ Issue {issue_key}: Using Priority '{severity_value}' as Severity fallback")
        else:
            severity_value = "Medium"

    # Normalize severity
    severity_map = {
        "Critical": "1 - Critical",
        "Blocker": "1 - Critical",
        "Highest": "1 - Critical",
        "High": "2 - High",
        "Major": "2 - High",
        "Medium": "3 - Medium",
        "Moderate": "3 - Medium",
        "Low": "4 - Low",
        "Minor": "4 - Low",
        "Trivial": "5 - Suggestion",
        "Lowest": "5 - Suggestion",
        "Suggestion": "5 - Suggestion",
        "Cosmetic": "5 - Suggestion"
    }
    severity = severity_map.get(str(severity_value).capitalize(), f"3 - {severity_value}")

    if idx <= 6:
        print(f"   Issue {issue_key}: Severity = '{severity_value}' → '{severity}'")

    # Tags
    labels = fields.get("labels", [])
    tags = ", ".join(labels) if isinstance(labels, list) else ""

    issue_url = f"{jira_url}/browse/{issue_key}"

    sheet.append([
        issue_key, issue_type, str(title),
        mapped_state, jira_status, assignee_name,
        tags, severity, issue_url
    ])

# Adjust widths
for col in range(1, 10):
    sheet.column_dimensions[get_column_letter(col)].width = 40

# Save
current_dir = os.path.dirname(os.path.abspath(__file__))
data_folder = os.path.join(current_dir, "data")
os.makedirs(data_folder, exist_ok=True)
save_path = os.path.join(data_folder, f"Jira {jira_project_key} Defects.xlsx")
wb.save(save_path)

elapsed_time = round(time.time() - start_time, 2)
print(f"\n✅ Excel file saved at: {save_path}")
print(f"⏱️ Total execution time: {elapsed_time} seconds")
print(f"📊 Total defects extracted: {len(all_issues)}")

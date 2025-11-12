import requests
from requests.auth import HTTPBasicAuth
import openpyxl
from openpyxl.utils import get_column_letter
from urllib.parse import quote
import os
import time

# Replace with your values
organization = "SobhaRealty"
project = "OneApp"
pat = os.getenv("DEVOPS_PAT")   # <-- set this in your environment
query_path = "My Queries/Smart-FM Replacement"
query_path_encoded = quote(query_path, safe='')

print("📄 Starting defects extraction...")
start_time = time.time()

# 1️⃣ Get saved query ID
print("📋 Fetching saved query...")
query_url = f"https://dev.azure.com/{organization}/{project}/_apis/wit/queries/{query_path_encoded}?api-version=7.0"
query_response = requests.get(query_url, auth=HTTPBasicAuth("", pat), timeout=30)

if query_response.status_code != 200:
    print(f"❌ Error fetching query: {query_response.status_code} - {query_response.text}")
    exit()

query_id = query_response.json().get("id")
if not query_id:
    print("❌ Could not get query ID from response.")
    exit()

print(f"✅ Query ID retrieved: {query_id}")

# 2️⃣ Run saved query to get work item IDs
print("🔍 Running saved query to fetch work items...")
run_query_url = f"https://dev.azure.com/{organization}/{project}/_apis/wit/wiql/{query_id}?api-version=7.0"
wiql_response = requests.get(run_query_url, auth=HTTPBasicAuth("", pat), timeout=30)

if wiql_response.status_code != 200:
    print(f"❌ Error running saved query: {wiql_response.status_code} - {wiql_response.text}")
    exit()

work_items = wiql_response.json().get("workItems", [])
ids = [str(item["id"]) for item in work_items]

if not ids:
    print("⚠️ No work items found.")
    exit()

print(f"✅ Found {len(ids)} work items. Fetching details...")

# 3️⃣ Create Excel
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Defects"
sheet.append(["ID", "Work Item Type", "Title", "State", "Assigned To", "Tags", "Environment", "Severity", "Issue Links"])

# 🔍 DEBUG: Track environment extraction
print("\n" + "="*80)
print("🔍 DEBUGGING ENVIRONMENT EXTRACTION")
print("="*80)

environment_stats = {
    "SIT": 0,
    "UAT": 0,
    "Both": 0,
    "None": 0,
    "samples": []
}

# 4️⃣ Fetch details for each work item with progress indicator
print("\n📥 Fetching work item details...")
for idx, work_id in enumerate(ids, start=1):
    if idx % 10 == 0 or idx == len(ids):
        print(f"   Progress: {idx}/{len(ids)} work items processed...")
    
    wi_url = f"https://dev.azure.com/{organization}/{project}/_apis/wit/workitems/{work_id}?api-version=7.0&$expand=Relations"
    
    try:
        wi_resp = requests.get(wi_url, auth=HTTPBasicAuth("", pat), timeout=30)
        
        if wi_resp.status_code != 200:
            print(f"⚠️ Failed to fetch work item {work_id}: {wi_resp.status_code}")
            continue

        wi_detail = wi_resp.json()
        fields = wi_detail.get("fields", {})

        # Direct DevOps URL for the defect
        issue_link = f"https://dev.azure.com/{organization}/{project}/_workitems/edit/{wi_detail.get('id', '')}"

        # Extract Tags
        tags = fields.get("System.Tags", "")
        
        # 🔍 DEBUG: Log ALL fields for first work item to find Environment field
        if idx == 1:
            print(f"\n{'='*80}")
            print(f"🔍 WORK ITEM {work_id} - FINDING ENVIRONMENT FIELD")
            print(f"{'='*80}")
            print("   All available fields:")
            for key in sorted(fields.keys()):
                value = fields.get(key)
                # Only show non-empty fields
                if value and value != "":
                    print(f"      {key} = '{value}'")
            print(f"{'='*80}\n")
        
        # 🔍 DEBUG: Log tags for first 5 items
        if idx <= 5:
            print(f"\n--- Work Item {work_id} ---")
            print(f"   Tags: '{tags}'")
        
        # Extract Environment - Try ALL possible field name patterns
        environment = ""
        
        # Search through ALL fields for anything containing "environment"
        for field_key, field_value in fields.items():
            if "environment" in field_key.lower() and field_value:
                environment = str(field_value).strip()
                if idx <= 5:
                    print(f"   ✅ Found Environment field: '{field_key}' = '{environment}'")
                break
        
        # If no dedicated field found, try parsing from tags as fallback
        if not environment and tags:
            tags_lower = tags.lower()
            tags_list = [t.strip() for t in str(tags).split(';')]
            
            # Look for SIT or UAT in individual tags
            for tag in tags_list:
                tag_upper = tag.upper()
                if tag_upper == "SIT":
                    environment = "SIT"
                    break
                elif tag_upper == "UAT":
                    environment = "UAT"
                    break
            
            if idx <= 5:
                print(f"   Parsing from tags: {tags_list}")
                if environment:
                    print(f"   ✅ Environment from tags: '{environment}'")
        
        # Update statistics
        if environment:
            env_upper = environment.upper()
            if "SIT" in env_upper and "UAT" in env_upper:
                environment_stats["Both"] += 1
            elif "SIT" in env_upper:
                environment_stats["SIT"] += 1
            elif "UAT" in env_upper:
                environment_stats["UAT"] += 1
            else:
                environment_stats["None"] += 1
        else:
            environment_stats["None"] += 1
        
        if idx <= 5:
            print(f"   Final Environment value: '{environment}'")
        
        # Store sample for debugging
        if len(environment_stats["samples"]) < 10:
            environment_stats["samples"].append({
                "id": work_id,
                "tags": tags,
                "environment": environment
            })

        # Extract data
        row_num = idx + 1
        sheet.cell(row=row_num, column=1, value=wi_detail.get("id", ""))
        sheet.cell(row=row_num, column=2, value=fields.get("System.WorkItemType", ""))
        sheet.cell(row=row_num, column=3, value=fields.get("System.Title", ""))
        sheet.cell(row=row_num, column=4, value=fields.get("System.State", ""))
        sheet.cell(row=row_num, column=5,
                   value=fields.get("System.AssignedTo", {}).get("displayName", "")
                   if fields.get("System.AssignedTo") else "")
        sheet.cell(row=row_num, column=6, value=tags)
        sheet.cell(row=row_num, column=7, value=environment)
        sheet.cell(row=row_num, column=8, value=fields.get("Microsoft.VSTS.Common.Severity", ""))
        sheet.cell(row=row_num, column=9, value=issue_link)
        
    except requests.exceptions.Timeout:
        print(f"⚠️ Timeout fetching work item {work_id}")
        continue
    except Exception as e:
        print(f"⚠️ Error processing work item {work_id}: {str(e)}")
        continue

# 🔍 DEBUG: Print environment extraction summary
print("\n" + "="*80)
print("📊 ENVIRONMENT EXTRACTION SUMMARY")
print("="*80)
print(f"Total work items processed: {len(ids)}")
print(f"Items with SIT only: {environment_stats['SIT']}")
print(f"Items with UAT only: {environment_stats['UAT']}")
print(f"Items with both SIT & UAT: {environment_stats['Both']}")
print(f"Items with no environment: {environment_stats['None']}")
print("\n📝 Sample entries:")
for sample in environment_stats["samples"]:
    print(f"   ID {sample['id']}: Tags='{sample['tags']}' → Environment='{sample['environment']}'")
print("="*80 + "\n")

# 5️⃣ Adjust column widths
for col in range(1, 10):
    sheet.column_dimensions[get_column_letter(col)].width = 40

# 6️⃣ Save Excel
current_dir = os.path.dirname(os.path.abspath(__file__))  # folder where this script lives
data_folder = os.path.join(current_dir, "data")           # 'data' folder inside repo
os.makedirs(data_folder, exist_ok=True)                   # create 'data' folder if missing

save_path = os.path.join(data_folder, "Smart FM Defects through Python.xlsx")
wb.save(save_path)

end_time = time.time()
elapsed_time = round(end_time - start_time, 2)

print(f"✅ Excel file saved at: {save_path}")
print(f"⏱️ Total execution time: {elapsed_time} seconds")
print(f"📊 Total defects extracted: {len(ids)}")

# 🔍 Additional debug info
print("\n💡 TIP: Check the Excel file 'Environment' column (Column G) to verify values")
print("💡 Expected values: 'SIT', 'UAT', 'SIT, UAT', or empty string")